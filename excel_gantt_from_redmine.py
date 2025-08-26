#
# Genarate an excel gantt chart from redmine issues.
#
# Issues' information to use:
#   issue_id, subject, assigned_to, start_date, due_date, closed_on, done_ratio
#

import datetime

import openpyxl
from openpyxl.formatting.rule import DataBarRule, FormulaRule
from openpyxl.styles import Border, Font, PatternFill, Side
from openpyxl.styles.alignment import Alignment
from openpyxl.utils.cell import get_column_letter
from redminelib import Redmine

from config import Config
from issue_dict import IssueData
from logging_helper import init_logger

# global variables
config = Config()
targeted_id = []  # Issue ID list those are extracted to process
registered_id = []  # Issue ID list those have been registered in excel already

logger = init_logger('excel_gantt_from_redmine', logfile_path='./log/excel_gantt_from_redmine.log')

def is_holiday(date: datetime.date) -> bool:
    """
    Check if the given date is a holiday or weekend.

    Args:
        date (datetime.date): The date to check.

    Returns:
        bool: True if the date is a holiday, saturday, or sunday, False otherwise.
    """

    w = date.weekday()  # day of week (0:monday - 6:sunday)
    if w == 5 or w == 6:
        return True
    
    for h in config.holidays:
        if date == h:
            return True

    return False

def set_title_row(ws) -> None:
    """
    Set title row and column width for gantt chart template.

    Args:
        ws (worksheet): excel worksheet
    """

    # set height of row
    # ws.row_dimensions[1].height = 40  # Title row

    # set column width
    ws.column_dimensions['A'].width =  8  # Task #
    ws.column_dimensions['B'].width = 50  # Subject
    ws.column_dimensions['C'].width = 16  # Assigned
    ws.column_dimensions['D'].width = 12  # Start Date
    ws.column_dimensions['E'].width = 12  # Due Date
    ws.column_dimensions['F'].width = 12  # Closed Date
    ws.column_dimensions['G'].width = 12  # Done Ratio

    fontname  = config.font_name

    ws.cell(1, 1).value = '#'
    ws.cell(1, 1).font = Font(name=fontname)
    ws.cell(1, 1).alignment = Alignment(horizontal='center', vertical='center')

    ws.cell(1, 2).value = 'Subject'
    ws.cell(1, 2).font = Font(name=fontname)
    ws.cell(1, 2).alignment = Alignment(horizontal='center', vertical='center')

    ws.cell(1, 3).value = 'Assigned'
    ws.cell(1, 3).font = Font(name=fontname)
    ws.cell(1, 3).alignment = Alignment(horizontal='center', vertical='center')

    ws.cell(1, 4).value = 'Start'
    ws.cell(1, 4).font = Font(name=fontname)
    ws.cell(1, 4).alignment = Alignment(horizontal='center', vertical='center')

    ws.cell(1, 5).value = 'Due'
    ws.cell(1, 5).font = Font(name=fontname)
    ws.cell(1, 5).alignment = Alignment(horizontal='center', vertical='center')

    ws.cell(1, 6).value = 'Closed'
    ws.cell(1, 6).font = Font(name=fontname)
    ws.cell(1, 6).alignment = Alignment(horizontal='center', vertical='center')

    ws.cell(1, 7).value = 'Done(%)'
    ws.cell(1, 7).font = Font(name=fontname)
    ws.cell(1, 7).alignment = Alignment(horizontal='center', vertical='center')

    # merge cells for title row
    ws.merge_cells('A1:A2')
    ws.merge_cells('B1:B2')
    ws.merge_cells('C1:C2')
    ws.merge_cells('D1:D2')
    ws.merge_cells('E1:E2')
    ws.merge_cells('F1:F2')
    ws.merge_cells('G1:G2')

def excel_set_gantt_chart_date(ws) -> None:
    """
    Set month and day for gantt chart in excel.

    Args:
        ws (worksheet): excel worksheet
    """

    start_gantt = config.start_date
    end_gantt = config.end_date
    fontname = config.font_name

    # fill color for holidays
    fillLightPink = PatternFill(patternType='solid', fgColor='ffccff')  # Light Pink

    column = 8  # H -
    d = start_gantt
    while d <= end_gantt:
        ws.column_dimensions[ get_column_letter(column) ].width = 4

        # Month
        if d == start_gantt or d.day == 1:
            ws.cell(1, column, d)
            ws.cell(1, column).number_format = 'mm'
            ws.cell(1, column).font = Font(name=fontname)
            ws.cell(1, column).alignment = Alignment(horizontal='center', vertical='center')

        # Day
        ws.cell(2, column, d)
        ws.cell(2, column).number_format = 'dd'
        ws.cell(2, column).font = Font(name=fontname)
        ws.cell(2, column).alignment = Alignment(horizontal='center', vertical='center')

        # fill on holiday column
        if is_holiday(ws.cell(2, column).value):
            ws.cell(2, column).fill = fillLightPink

        d += datetime.timedelta(days=1)
        column += 1

def write_issue(ws, issue_data, indent: int, row: int) -> int:
    """
    Write issue information to the excel worksheet.

    Args:
        ws (worksheet): excel worksheet
        issue (IssueData): issue object
        indent (int): Indentation level for the issue
        row (int): Current row number in the worksheet

    Returns:
        int: Updated row number after writing the issue
    """

    fontname = config.font_name
    linkURLbase = config.link_url

    id = issue_data.id
    if id in registered_id:
        return row
    else:
        registered_id.append(id)

    ws.cell(row, 1).number_format = openpyxl.styles.numbers.FORMAT_GENERAL
    ws.cell(row, 1).value = issue_data.id
    ws.cell(row, 1).font = Font(name=fontname, color='0563C1', underline='single')
    ws.cell(row, 1).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row, 1).hyperlink = f'{linkURLbase}{issue_data.id}'

    ws.cell(row, 2).number_format = openpyxl.styles.numbers.FORMAT_GENERAL
    ws.cell(row, 2).value = issue_data.subject
    ws.cell(row, 2).font = Font(name=fontname)
    ws.cell(row, 2).alignment = Alignment(indent=indent, vertical='center')
    if id not in targeted_id:
        # This is not a target issue in this filter, it should be a parent issue of one of the target issue
        ws.cell(row, 2).fill = PatternFill(patternType='solid', fgColor='D9D9D9')

    ws.cell(row, 3).number_format = openpyxl.styles.numbers.FORMAT_GENERAL
    ws.cell(row, 3).value = issue_data.assigned_to if issue_data.assigned_to is not None else ''
    ws.cell(row, 3).font = Font(name=fontname)
    ws.cell(row, 3).alignment = Alignment(horizontal='center', vertical='center')

    ws.cell(row, 4).number_format = 'yyyy/mm/dd'
    ws.cell(row, 4).value = issue_data.start_date if issue_data.start_date is not None else ''
    ws.cell(row, 4).font = Font(name=fontname)
    ws.cell(row, 4).alignment = Alignment(horizontal='center', vertical='center')

    ws.cell(row, 5).number_format = 'yyyy/mm/dd'
    ws.cell(row, 5).value = issue_data.due_date if issue_data.due_date is not None else ''
    ws.cell(row, 5).font = Font(name=fontname)
    ws.cell(row, 5).alignment = Alignment(horizontal='center', vertical='center')

    ws.cell(row, 6).number_format = 'yyyy/mm/dd'
    ws.cell(row, 6).value = issue_data.closed_on if issue_data.closed_on is not None else ''
    ws.cell(row, 6).font = Font(name=fontname)
    ws.cell(row, 6).alignment = Alignment(horizontal='center', vertical='center')

    ws.cell(row, 7).number_format = openpyxl.styles.numbers.FORMAT_PERCENTAGE
    ws.cell(row, 7).value = issue_data.done_ratio / 100 if issue_data.done_ratio is not None else ''
    ws.cell(row, 7).font = Font(name=fontname)
    ws.cell(row, 7).alignment = Alignment(horizontal='center', vertical='center')

    # If the issue is closed, set the done ratio to 100%
    if issue_data.closed_on is not None:
        ws.cell(row, 7).value = 1.0  # 100% complete

    return row+1

def set_conditional_format(ws, min_row: int, max_row: int) -> None:
    """
    Set conditional formatting for gantt chart template.

    Args:
        ws (worksheet): excel worksheet
        min_row (int): minimum row number for gantt chart
        max_row (int): maximum row number for gantt chart
    """

    start_gantt = config.start_date
    end_gantt = config.end_date

    # progress bar : F
    r1 = DataBarRule(start_type='num', start_value=0, end_type='num', end_value=1, color='31869B', showValue=True, minLength=0, maxLength=100)
    range = f'$G${min_row}:$G${max_row}'
    ws.conditional_formatting.add(range, r1)

    # gantt chart : H -
    # count the number of date columns and find the last column
    start_gantt_column = 8  # H -
    days = nColumnGantt = end_gantt - start_gantt
    nColumnGantt = days.days
    end_gantt_column = start_gantt_column + nColumnGantt

    # condition 1 : completed part considering progress percentage
    c1 = '=AND( $D3<=H$2, H$2<=ROUNDDOWN( ($E3-$D3+1)*$G3, 0 )+$D3-1 )'
    # condition 2 : uncompleted part considering progress percentage
    c2 = '=AND( $D3<=H$2, H$2<=$E3 )'
    # condition 3 : task for future
    c3 = '=AND( $D3<=H$2, H$2<=$E3, TODAY()<H$2 )'
    # condition 4 : today
    c4 = '=AND( H$2=TODAY() )'
    # condition 5 : overdue (due cells)
    c5 = '=AND( $E3<>"", $E3<TODAY(), $G3<1 )'

    # fromat 1 : fill completed part
    f1 = PatternFill(patternType='solid', bgColor='8888ff')
    # formay 2 : fill uncompleted part
    f2 = PatternFill(patternType='solid', bgColor='ff8888')
    # format 3 : future task
    f3 = PatternFill(patternType='solid', bgColor='cccccc')
    # format 4 : today
    f4 = PatternFill(patternType='lightGray', fgColor='31869b')
    # format 5 : overdue (due cells)
    f5 = PatternFill(patternType='solid', bgColor='ffff88')

    # combine conditions and formats
    r1 = FormulaRule(formula=[c1] , stopIfTrue=None, fill=f1)
    r2 = FormulaRule(formula=[c2] , stopIfTrue=None, fill=f2)
    r3 = FormulaRule(formula=[c3] , stopIfTrue=None, fill=f3)
    r4 = FormulaRule(formula=[c4] , stopIfTrue=None, fill=f4)
    r5 = FormulaRule(formula=[c5] , stopIfTrue=None, fill=f5)

    # set conditional format
    start_cell = f'${'H'}${min_row}'
    end_cell   = f'${get_column_letter(end_gantt_column)}${max_row}'
    cells      = start_cell + ':' + end_cell
    ws.conditional_formatting.add(cells, r1)
    ws.conditional_formatting.add(cells, r2)
    ws.conditional_formatting.add(cells, r3)
    start_cell = f'${'H'}${min_row-1}' # (-1) because including month row
    end_cell   = f'${get_column_letter(end_gantt_column)}${max_row}'
    cells      = start_cell + ':' + end_cell
    ws.conditional_formatting.add(cells, r4)
    start_cell = f'${'E'}${min_row}' # from due date column
    end_cell   = f'${'E'}${max_row}' # to due date column
    cells      = start_cell + ':' + end_cell
    ws.conditional_formatting.add(cells, r5)

    # fill holiday cells
    r = min_row
    fillLightPink = PatternFill(patternType='solid', fgColor='ffdcff')  # Light Pink
    side = Side(style='thin', color='aaaaaa')
    border = Border(top=side, bottom=side, left=side, right=side)
    while r <= max_row:
        # set_task_format(ws, r)
        c = start_gantt_column
        while c <= end_gantt_column:
            v = ws.cell(2, c).value
            if is_holiday(v):
                ws.cell(r, c).fill = fillLightPink
            ws.cell(r, c).border = border  # set border line to all cells in gantt chart area
            c += 1
        r += 1

def get_filter_issues(redmine, filter: dict) -> (dict|None):
    global targeted_id
    try:
        # Search filter conditions
        issues = redmine.issue.filter(**filter)

        issues_dict = dict()
        for issue in issues:
            issue_data = IssueData()
            issue_data.id          = issue.id
            issue_data.subject     = issue.subject
            issue_data.assigned_to = issue.assigned_to.name if hasattr(issue, 'assigned_to') else None
            issue_data.start_date  = issue.start_date if hasattr(issue, 'start_date') else None
            issue_data.due_date    = issue.due_date if hasattr(issue, 'due_date') else None
            issue_data.closed_on   = issue.closed_on if hasattr(issue, 'closed_on') else None
            issue_data.done_ratio  = issue.done_ratio if hasattr(issue, 'done_ratio') else None
            issue_data.parent_id   = issue.parent.id if hasattr(issue, 'parent') else None
            issues_dict[issue.id] = issue_data

            targeted_id.append(issue.id)

        return issues_dict
    except Exception as e:
        logger.error(f'Redmine error : {e}')
        return None

def get_ancestor_issues(redmine, issues_dict: dict) -> (dict|None):
    try:
        ancestors_dict = dict()

        for id, issue_data in issues_dict.items():
            parent_id = issue_data.parent_id
            if parent_id is not None:
                if parent_id in issues_dict:
                    issues_dict[parent_id].children_id.append(id)
                elif parent_id in ancestors_dict:
                    ancestors_dict[parent_id].children_id.append(id)
                else:
                    while parent_id is not None:
                        parent_issue = redmine.issue.get(parent_id)
                        parent_data = IssueData()
                        parent_data.id          = parent_issue.id
                        parent_data.subject     = parent_issue.subject
                        parent_data.assigned_to = parent_issue.assigned_to.name if hasattr(parent_issue, 'assigned_to') else None
                        parent_data.start_date  = parent_issue.start_date if hasattr(parent_issue, 'start_date') else None
                        parent_data.due_date    = parent_issue.due_date if hasattr(parent_issue, 'due_date') else None
                        parent_data.closed_on   = parent_issue.closed_on if hasattr(parent_issue, 'closed_on') else None
                        parent_data.done_ratio  = parent_issue.done_ratio if hasattr(parent_issue, 'done_ratio') else None
                        parent_data.parent_id   = parent_issue.parent.id if hasattr(parent_issue, 'parent') else None
                        ancestors_dict[parent_id] = parent_data

                        ancestors_dict[parent_id].children_id.append(id)
                        parent_id = parent_data.parent_id
        return ancestors_dict
    except Exception as e:
        logger.error(f'Redmine error : {e}')
        return None

def get_topmost_id(id, issues_dict) -> int:
    topmost_id = id
    while issues_dict[topmost_id].parent_id is not None:
        topmost_id = issues_dict[topmost_id].parent_id
    return topmost_id

def main() -> None:
    redmine = Redmine(config.url, username=config.username, password=config.password)

    filter = {
        'project_id': config.project_name,
    }
    if config.sort:
        filter['sort'] = config.sort
    if config.issue_id:
        filter['issue_id'] = config.issue_id
    if config.query_id:
        filter['query_id'] = config.query_id
    if config.parent_id:
        filter['parent_id'] = config.parent_id
    if config.tracker_id:
        filter['tracker_id'] = config.tracker_id
    if config.status_id:
        filter['status_id'] = config.status_id
    if config.author_id:
        filter['author_id'] = config.author_id
    if config.assigned_to_id:
        filter['assigned_to_id'] = config.assigned_to_id
    if config.fixed_version_id:
        filter['fixed_version_id'] = config.fixed_version_id

    # get issues according to the specified filter condition
    issues_dict = get_filter_issues(redmine, filter)
    if issues_dict is None:
        return

    # Number of items that match the search criteria
    total = len(issues_dict)
    if total == 0:
        logger.info('No issues found with the specified filter.')
        return
    logger.info(f'Total found issues : {total}')

    # get ancestor(not only parent) issues associated with the target issues
    ancestors_dict = get_ancestor_issues(redmine, issues_dict)

    # marge ancestor dict to issues dict
    if ancestors_dict is not None:
        issues_dict.update(ancestors_dict)
        ancestors_dict.clear()
        # re-calc total issues to process
        total = len(issues_dict)

    wb = openpyxl.Workbook()
    ws = wb.worksheets[0]

    # Tab title
    if config.tab_title:
        ws.title = config.tab_title

    # Title row
    set_title_row(ws)

    # Set filter
    # ws.dimensions applies filter to all columns with data
    # By setting filter before entering dates, we can set filter only on non-date columns
    # If set after entering dates, filter will also be set on date columns, so it's important to do it here
    # ws.auto_filter.ref = ws.dimensions

    # Date row for gantt chart
    excel_set_gantt_chart_date(ws)

    import datetime
    t0 = datetime.datetime.now()

    row = 3
    progress = 0

    global registered_id
    for id, issue_data in issues_dict.items():
        def display_progress(progress, total):
            p = int(progress*100/total)
            print(f'\r [ {p:2}% ] done.', end='')

        if issue_data.id in registered_id:
            continue

        topmost_id = get_topmost_id(id, issues_dict)
        row = write_issue(ws, issues_dict[topmost_id], 0, row)
        progress += 1
        display_progress(progress, total)

        if topmost_id == id:
            continue
        else:
            # process children tree recursively
            def process_children_tree(parent_id: int, indent: int, row: int) -> int:
                nonlocal progress
                children_list = issues_dict[parent_id].children_id
                for child_id in children_list:
                    row = write_issue(ws, issues_dict[child_id], indent, row)
                    progress += 1
                    display_progress(progress, total)
                    if issues_dict[child_id].children_id:
                        row = process_children_tree(child_id, indent+1, row)
                return row
            row = process_children_tree(topmost_id, 1, row)

    t1 = datetime.datetime.now()
    logger.info(f'Total process time : {t1-t0}')

    # Freeze window panes
    ws.freeze_panes = 'H3'
    # Set filter
    ws.auto_filter.ref = f'A2:G{row-1}'

    # Conditional formatting
    set_conditional_format(ws, 3, row-1)

    while True:
        print(" Input file name (It doesn't need '.xlsx' extention.) : ", end='')
        f = input()
        try:
            wb.save(f'.\\{f}.xlsx')
            break
        except Exception:
            logger.error(f" Error : Can't save to '{f}.xlsx'.")
            print(' Do you want to try again? [_/n] : ', end='')
            yn = input().upper()
            if yn == 'N':
                break 

if __name__ == '__main__':
    if config.load_config_from_toml():
        config.user_account()
        main()

